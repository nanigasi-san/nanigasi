#シート数・シート名の取得
import openpyxl

book = openpyxl.load_workbook("library/use_file/openpyxl/xlsx/test_book.xlsx")

print("-"*25)
print(len(book.sheetnames))

print("-"*25)
for name in book.sheetnames:# book.get_sheet_names() ではエラーが出る
    print(name)
